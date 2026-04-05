"""
Excel export — filtered deal list + summary tables in a formatted workbook.
Multi-sheet: Deals, Valuation Summary, Sponsor Rankings.
Uses openpyxl for formatting. data_origin column always preserved.
"""
import os
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from ma.db import queries
from ma.analytics import valuation, sponsor_intel


def export_deals_excel(filters: dict = None, config: dict = None, filename: str = None) -> str:
    """
    Export filtered deals to Excel with multiple formatted sheets.
    Returns the full output filepath.
    """
    exp_cfg = (config or {}).get("export", {}).get("excel", {})
    output_dir = exp_cfg.get("output_dir", "outputs")
    os.makedirs(output_dir, exist_ok=True)

    if not filename:
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"ma_deals_{ts}.xlsx"

    filepath = os.path.join(output_dir, filename)

    deals_df = queries.get_all_deals(filters)
    val_stats = valuation.sector_valuation_stats(filters)
    sponsor_df = sponsor_intel.sponsor_rankings(filters, top_n=20)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Sheet 1: Deals
        deals_df.to_excel(writer, sheet_name="Deals", index=False)
        _format_sheet(writer.sheets["Deals"], deals_df)

        # Sheet 2: Valuation Summary (if enabled)
        if exp_cfg.get("include_valuation_sheet", True) and not val_stats.empty:
            val_stats.to_excel(writer, sheet_name="Valuation by Sector", index=False)
            _format_sheet(writer.sheets["Valuation by Sector"], val_stats)

        # Sheet 3: Sponsor Rankings (if enabled)
        if exp_cfg.get("include_summary_sheet", True) and not sponsor_df.empty:
            sponsor_df.to_excel(writer, sheet_name="Sponsor Rankings", index=False)
            _format_sheet(writer.sheets["Sponsor Rankings"], sponsor_df)

    return filepath


def _format_sheet(ws, df: pd.DataFrame) -> None:
    """
    Apply basic Excel formatting:
    - Header row: bold, dark background, white text
    - Column widths: auto-fit (capped at 50)
    """
    header_fill = PatternFill(start_color="1C2833", end_color="1C2833", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)), df[col].astype(str).str.len().max() if not df.empty else 10)
        width = min(max_len + 4, 50)
        ws.column_dimensions[get_column_letter(i)].width = width
